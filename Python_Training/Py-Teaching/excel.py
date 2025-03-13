#!/usr/bin/env python3
# -*- coding: utf-8 -*-

def getFillCell(cell) -> str:
	return str(cell.fill.bgColor.rgb) if str(cell.fill.bgColor.rgb) == str(cell.fill.fgColor.rgb) else str(str(cell.fill.bgColor.rgb) if len(str(cell.fill.bgColor.rgb)) <= 10 else str(cell.fill.fgColor.rgb))

def main():
	# pip install openpyxl
	import openpyxl
	wb = openpyxl.load_workbook(filename='./Услуги-2024.xlsm', data_only=True, read_only=True, keep_vba=True)
	#print(wb.sheetnames)
	#print(wb["Январь"].title)
	#anotherSheet = wb.active
	#print(anotherSheet.title)
	#print(wb["Январь"].max_row, wb["Январь"].max_column)
	#print(wb["Февраль"].max_row, wb["Февраль"].max_column)
	#print(wb["Январь"].cell(5, 1).value)
	#print(wb["Январь"].cell(12, 1).value)
	#print(wb["Февраль"].cell(5, 1).value)
	#print(wb["Февраль"].cell(12, 1).value)
	#wb.save('Услуги-2023.xlsm')
	#print(dir(wb))
	
	#oncell = wb["Сентябрь"].cell(12, 7)
	#print(oncell.value, getFillCell(oncell))
	#oncell = wb["Сентябрь"].cell(13, 7)
	#print(oncell.value, getFillCell(oncell))
	
	#oncell = wb["Сентябрь"].cell(12, 4)
	#print(oncell.value)
	#oncell = wb["Сентябрь"].cell(12, 5)
	#print(oncell.value)
	
	pass

if __name__ == '__main__':
	main()
